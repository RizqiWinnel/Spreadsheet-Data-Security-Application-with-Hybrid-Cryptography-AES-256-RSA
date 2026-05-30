# backend_file_processor.py
import io
import math
from PIL import Image
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl import load_workbook
import backend_aes
import backend_rsa

def parse_cell_value(value_str):
    """
    Mengembalikan string hasil dekripsi ke tipe data aslinya (int, float, atau tetap str)
    agar bisa dibaca dengan benar oleh Excel.
    """
    # 1. Cek apakah ini formula Excel
    if value_str.startswith('='):
        return value_str
        
    # 2. Coba konversi ke Integer (bilangan bulat)
    try:
        return int(value_str)
    except ValueError:
        pass 
        
    # 3. Coba konversi ke Float (bilangan desimal)
    try:
        return float(value_str)
    except ValueError:
        pass
        
    # 4. Jika semua gagal, berarti ini memang teks murni (string)
    return value_str

# ==========================================
# 1. LOGIKA MANIPULASI PIKSEL GAMBAR
# ==========================================
def image_encrypt(image_bytes, aes_key):
    try:
        # Buka gambar
        img = Image.open(io.BytesIO(image_bytes))
        mode = img.mode
        
        # Penanganan Ruang Warna Fleksibel
        if mode not in ['RGB', 'RGBA', 'L', '1']:
            if 'transparency' in img.info or 'A' in mode:
                img = img.convert('RGBA')
                mode = 'RGBA'
            else:
                img = img.convert('RGB')
                mode = 'RGB'
            
        if mode == '1':
            img_to_process = img.convert('L')
        else:
            img_to_process = img
            
        orig_w, orig_h = img_to_process.size
        raw_pixel_bytes = img_to_process.tobytes()
        
        # 3. Sisipkan Metadata (Dimensi & Mode)
        metadata = f"{orig_w},{orig_h},{mode}|".encode('utf-8')
        data_to_encrypt = metadata + raw_pixel_bytes
        
        # 4. Enkripsi murni tingkat piksel menggunakan mode ECB
        encrypted_bytes = backend_aes.encrypt_aes(data_to_encrypt, aes_key)
        if encrypted_bytes is None: return None
            
        # 5. Header Panjang Ciphertext untuk keamanan pemotongan
        cipher_len = len(encrypted_bytes)
        header_len = f"{cipher_len}|".encode('utf-8')
        full_payload = header_len + encrypted_bytes
        
        # 6. Render array piksel terenkripsi menjadi kanvas persegi (RGB Base)
        sisa = len(full_payload) % 3
        if sisa != 0:
            full_payload += b'\x00' * (3 - sisa)
            
        total_pixels = len(full_payload) // 3
        new_w = math.ceil(math.sqrt(total_pixels))
        new_h = math.ceil(total_pixels / new_w)
        
        req_bytes = new_w * new_h * 3
        if len(full_payload) < req_bytes:
            full_payload += b'\x00' * (req_bytes - len(full_payload))
            
        enc_img = Image.frombytes('RGB', (new_w, new_h), full_payload)
        output_io = io.BytesIO()
        enc_img.save(output_io, format='PNG') # Simpan
        return output_io.getvalue()
        
    except Exception as e:
        print(f"Error Enkripsi Piksel: {e}")
        return None

def image_decrypt(encrypted_image_bytes, aes_key):
    try:
        # 1. Buka gambar ciphertext persegi
        enc_img = Image.open(io.BytesIO(encrypted_image_bytes)).convert('RGB')
        raw_bytes = enc_img.tobytes()
        
        # 2. Baca Header Panjang
        sep_idx = raw_bytes.find(b'|')
        if sep_idx == -1: return None
            
        cipher_len = int(raw_bytes[:sep_idx].decode('utf-8'))
        encrypted_bytes = raw_bytes[sep_idx + 1 : sep_idx + 1 + cipher_len]
        
        # 3. Dekripsi AES
        decrypted_data = backend_aes.decrypt_aes(encrypted_bytes, aes_key)
        if decrypted_data is None: return None
            
        # 4. Baca Metadata (Dimensi & Mode Asli)
        sep_idx2 = decrypted_data.find(b'|')
        metadata_str = decrypted_data[:sep_idx2].decode('utf-8')
        orig_w_str, orig_h_str, orig_mode = metadata_str.split(',')
        orig_w, orig_h = int(orig_w_str), int(orig_h_str)
        
        # 5. Ekstrak data piksel asli tanpa byte padding
        pixel_bytes_asli = decrypted_data[sep_idx2 + 1 :]
        
        # 6. Rekonstruksi Matriks Gambar berdasarkan Mode Warnanya
        if orig_mode == 'RGBA':
            expected_length = orig_w * orig_h * 4
            pixel_bytes_asli = pixel_bytes_asli[:expected_length]
            orig_img = Image.frombytes('RGBA', (orig_w, orig_h), pixel_bytes_asli)
            
        elif orig_mode == '1':
            expected_length = orig_w * orig_h
            pixel_bytes_asli = pixel_bytes_asli[:expected_length]
            orig_img = Image.frombytes('L', (orig_w, orig_h), pixel_bytes_asli).convert('1')
            
        elif orig_mode == 'L':
            expected_length = orig_w * orig_h
            pixel_bytes_asli = pixel_bytes_asli[:expected_length]
            orig_img = Image.frombytes('L', (orig_w, orig_h), pixel_bytes_asli)
            
        else: # Default ke RGB
            expected_length = orig_w * orig_h * 3
            pixel_bytes_asli = pixel_bytes_asli[:expected_length]
            orig_img = Image.frombytes('RGB', (orig_w, orig_h), pixel_bytes_asli)
            
        # 7. Simpan hasil rekonstruksi untuk dimasukkan ke Excel
        output_io = io.BytesIO()
        orig_img.save(output_io, format='PNG')
        return output_io.getvalue()
        
    except Exception as e:
        print(f"Error Dekripsi Piksel: {e}")
        return None

# ==========================================
# 2. LOGIKA ITERASI EXCEL (ENKRIPSI & DEKRIPSI)
# ==========================================

def proses_enkripsi_dokumen(path_excel, public_key_rsa_str, aes_key_bytes):
    """
    Membaca Excel, mengenkripsi sel dan gambar in-place,
    menyimpan Excel baru, dan menghasilkan Kunci AES terenkripsi RSA (.txt).
    """
    print(f"--- Memulai Enkripsi File: {path_excel} ---")
    try:
        # 1. Kunci AES
        aes_key = aes_key_bytes
        
        # 2. Buka Workbook Excel
        wb = load_workbook(path_excel)
        
        # 3. Iterasi setiap Sheet
        for sheet in wb.worksheets:
            print(f"Memproses Sheet: {sheet.title}")
            
            # --- A. Proses Sel (Teks & Formula) ---
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # Ubah isi sel jadi string (termasuk formula)
                        teks_asli = str(cell.value).encode('utf-8')
                        
                        # Enkripsi dengan AES
                        cipher_bytes = backend_aes.encrypt_aes(teks_asli, aes_key)
                        
                        # Ubah ke Hex agar aman ditulis di sel Excel
                        cipher_hex = cipher_bytes.hex().upper()
                        
                        # Timpa sel dengan penanda
                        cell.value = cipher_hex
            
            # --- B. Proses Gambar (In-Place) ---
            if sheet._images:
                gambar_asli_list = list(sheet._images) # Buat salinan
                sheet._images.clear() # Kosongkan gambar dari sheet
                
                for img in gambar_asli_list:
                    # Ambil properti gambar
                    orig_width = img.width
                    orig_height = img.height
                    orig_anchor = img.anchor
                    
                    # Ambil byte data gambar
                    # Di openpyxl, data mentah gambar biasanya ada di img.ref atau dipanggil via _data()
                    try:
                        img_bytes = img._data() 
                    except TypeError:
                        img_bytes = img._data # Menyesuaikan versi openpyxl
                    
                    # Proses piksel (fungsi yang sebelumnya kita buat)
                    encrypted_img_bytes = image_encrypt(img_bytes, aes_key)
                    
                    if encrypted_img_bytes:
                        # Buat objek gambar openpyxl baru dari memori
                        virtual_file = io.BytesIO(encrypted_img_bytes)
                        new_img = OpenpyxlImage(virtual_file)
                        
                        # Kembalikan ukuran dan posisi semula
                        new_img.width = orig_width
                        new_img.height = orig_height
                        new_img.anchor = orig_anchor
                        
                        # Sisipkan kembali ke sheet
                        sheet.add_image(new_img)
        
        # 5. Amankan Kunci AES dengan Kunci Publik RSA
        status, cipher_key_str = backend_rsa.encrypt_rsa(aes_key, public_key_rsa_str)
        
        if not status:
            return False, f"Gagal mengenkripsi kunci AES: {cipher_key_str}"
            
        return True, (wb, cipher_key_str)
        
    except Exception as e:
        return False, f"Terjadi kesalahan saat enkripsi Excel: {str(e)}"

# --- FUNGSI DEKRIPSI ---
def proses_dekripsi_dokumen(path_excel_enc, private_key_rsa_str, cipher_key_aes_str):
    """
    Membaca file Excel terenkripsi, memulihkan Kunci AES dari file .txt menggunakan RSA,
    lalu mendekripsi seluruh sel dan gambar di dalam Excel.
    """
    print(f"--- Memulai Dekripsi File: {path_excel_enc} ---")
    try:
        # 1. Dekripsi Kunci AES menggunakan RSA    
        status, aes_key = backend_rsa.decrypt_rsa(cipher_key_aes_str, private_key_rsa_str)
        if not status:
            return False, f"Gagal mendekripsi kunci AES: {aes_key}"
            
        # 2. Buka Workbook Terenkripsi
        wb = load_workbook(path_excel_enc)
        
        # 3. Iterasi Sheet
        for sheet in wb.worksheets:
            print(f"Mendekripsi Sheet: {sheet.title}")
            
            # --- A. Dekripsi Sel (teks & formula) ---
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cipher_hex = str(cell.value)
                        cipher_bytes = bytes.fromhex(cipher_hex)
                        
                        # Dekripsi AES
                        plain_bytes = backend_aes.decrypt_aes(cipher_bytes, aes_key)
                        plain_str = plain_bytes.decode('utf-8')
                        
                        # Kembalikan ke tipe data asli (int/float/formula)
                        cell.value = parse_cell_value(plain_str)
            
            # --- B. Dekripsi Gambar ---
            if sheet._images:
                gambar_enc_list = list(sheet._images)
                sheet._images.clear()
                
                for img in gambar_enc_list:
                    # 1. Ambil posisi DAN Dimensi Skala Skripsi
                    orig_anchor = img.anchor
                    orig_width = img.width
                    orig_height = img.height
                    try:
                        img_bytes = img._data()
                    except TypeError:
                        img_bytes = img._data
                        
                    # Panggil fungsi dekripsi piksel
                    decrypted_img_bytes = image_decrypt(img_bytes, aes_key)
                    
                    if decrypted_img_bytes:
                        virtual_file = io.BytesIO(decrypted_img_bytes)
                        new_img = OpenpyxlImage(virtual_file)
                        
                        # 2. Kembalikan posisi DAN Dimensi Skala
                        new_img.anchor = orig_anchor
                        new_img.width = orig_width
                        new_img.height = orig_height

                        sheet.add_image(new_img)
                        
        # KEMBALIKAN OBJEK WORKBOOK DAN KUNCI AES
        return True, (wb, aes_key.decode('utf-8'))
        
    except Exception as e:
        return False, f"Terjadi kesalahan saat dekripsi Excel: {str(e)}"